# excel_manager.py

from openpyxl import Workbook, load_workbook
import os

EXCEL_FILE = "attendance.xlsx"

BASE_HEADERS = ["Roll No", "Name", "Presents", "Absents", "Percentage"]

def update_attendance_sheet(roll_no, name, date, status):
    # Create file if not exists
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(BASE_HEADERS)
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # 🔹 Ensure date column exists
    headers = [cell.value for cell in ws[1]]
    if date not in headers:
        ws.cell(row=1, column=len(headers) + 1).value = date
        headers.append(date)

    date_col = headers.index(date) + 1

    student_row = None

    # 🔹 Find student row
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == roll_no:
            student_row = row
            break

    # 🔹 Create student row if not found
    if student_row is None:
        student_row = ws.max_row + 1
        ws.cell(row=student_row, column=1).value = roll_no
        ws.cell(row=student_row, column=2).value = name
        ws.cell(row=student_row, column=3).value = 0  # Presents
        ws.cell(row=student_row, column=4).value = 0  # Absents
        ws.cell(row=student_row, column=5).value = 0  # Percentage

    # 🔹 Avoid duplicate marking
    if ws.cell(row=student_row, column=date_col).value:
        wb.close()
        return

    # 🔹 Mark attendance
    ws.cell(row=student_row, column=date_col).value = status

    presents = ws.cell(row=student_row, column=3).value
    absents = ws.cell(row=student_row, column=4).value

    presents = int(presents)
    absents = int(absents)

    if status.lower() == "Present":
        presents += 1
    else:
        absents += 1

    ws.cell(row=student_row, column=3).value = presents
    ws.cell(row=student_row, column=4).value = absents

    total = presents + absents
    percentage = round((presents / total) * 100, 2) if total > 0 else 0
    ws.cell(row=student_row, column=5).value = percentage

    wb.save(EXCEL_FILE)
    wb.close()
